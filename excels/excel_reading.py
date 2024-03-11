import openpyxl

wb = openpyxl.load_workbook("store.xlsx")
print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

sheet = wb["Products"]
sheet = wb.active

b2_cell = sheet["B2"]
print(b2_cell.value)
print(sheet.cell(row=2, column=2).value)